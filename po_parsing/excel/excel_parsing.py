# pip install pandas openpyxl
import pandas as pd
from openpyxl import load_workbook

def extract_text_and_tables_in_order(file_path):
    # 눈에 보기 좋게 출력하기 위해 load_workbook 사용
    wb = load_workbook(file_path)

    # 엑셀 파일에서 모든 시트를 읽기
    dfs = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')

    sheet_texts = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_text = f"--- Sheet: {sheet_name} ---\n"
        
        for row in ws.iter_rows(values_only=True):
            row_values = [str(cell).strip() if cell is not None else "" for cell in row]
            # 모든 셀이 빈 경우는 건너뛰기 (줄 정리용)
            if any(cell for cell in row_values):
                sheet_text += "\t".join(row_values) + "\n"

        sheet_texts.append(sheet_text)

    full_content = "\n\n".join(sheet_texts)
    return(full_content)
